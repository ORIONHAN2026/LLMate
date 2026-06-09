#!/usr/bin/env python3
"""
SKU重复检查脚本（一品多码检测）
功能：
  1. 读取数据总表：Sheet1（商品原始表）+ Sheet2（动态商品池）
  2. 品牌模糊匹配（子串包含）+ 名称混合匹配，若匹配到但编码不同 → 判定为「一品多码」
  3. 将动态商品池中一品多码的行以红色标注
  4. 输出 JSON 格式对比报告

判定规则：
  - 品牌匹配（精确或子串）+ 名称匹配（精确/子串/关键词重叠）+ 编码相同 → 同一商品，不标红
  - 品牌匹配（精确或子串）+ 名称匹配（精确/子串/关键词重叠）+ 编码不同 → 一品多码，标红
  - 品牌不匹配 或 名称不匹配 → 不同商品，不标红

名称匹配规则（混合策略）：
  1. 子串包含：标准化后一个名称完全包含另一个 → 匹配
  2. 关键词重叠：子串不命中时，拆分关键词集合并算重叠系数（≥60%即匹配）
  例如：「石墨盘根 10*10mm」包含于「森耐德 石墨盘根 10*10mm 5kg/卷」→ 子串匹配
  例如：「HP 硒鼓 CF280A 黑色」与「惠普 HP 硒鼓 CF280A/80a 黑色」→ 关键词重叠匹配

表头结构（固定）：
  Sheet1（商品原始表）: 商品编码 | 自定义名称 | 品牌
  Sheet2（动态商品池）: 名称       | 品牌       | 编码 | 单位
"""

import sys
import os
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# ── 配置 ──────────────────────────────────────────────
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)

# 固定列配置
# Sheet1（商品原始表）列映射
MASTER_COLS = {
    "编码": 0,
    "名称": 1,   # 实际列名: 自定义名称
    "品牌": 2,
}
# Sheet2（动态商品池）列映射
POOL_COLS = {
    "名称": 0,
    "品牌": 1,
    "编码": 2,
    "单位": 3,
}


def normalize_value(val):
    """标准化单元格值：去首尾空格，转字符串，空值返回空字符串。"""
    if val is None:
        return ""
    return str(val).strip()


def brands_match(brand_a, brand_b):
    """
    判断两个品牌是否指向同一品牌（子串包含匹配）。
    去除括号及其内容、空格后，一个品牌完全包含另一个即视为匹配。
    例如：「正泰」与「正泰（CHNT）」→ 匹配
    """
    def normalize_brand(b):
        if not b:
            return ""
        b = re.sub(r'\([^)]*\)', '', b)
        b = re.sub(r'（[^）]*）', '', b)
        b = re.sub(r'\s+', '', b)
        return b.lower()

    na = normalize_brand(brand_a)
    nb = normalize_brand(brand_b)
    if not na or not nb:
        return False
    return na == nb or na in nb or nb in na


def normalize_name_for_compare(text):
    """
    标准化名称用于子串包含比较。
    - 去除所有空格
    - 去除括号及其内容（中英文括号）
    - 转小写
    返回标准化后的字符串，若为空则返回空字符串。
    """
    if not text:
        return ""
    text = re.sub(r'\s+', '', text)
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.sub(r'（[^）]*）', '', text)
    return text.lower()


def tokenize_name(name):
    """
    将名称拆分为关键词集合，用于重叠度计算。
    - 按常见分隔符/标点拆分
    - 过滤纯数字/纯符号的 token
    """
    if not name:
        return set()
    # 用分隔符拆分
    # 按分隔符/标点/空白拆分
    delimiters = re.compile(r'[\s!"#%&\'()*+,\-./:;<=>?@\[\\\]^_`{|}~\uff00-\uffef\u3000-\u303f\uff01-\uff5e]+')
    raw_tokens = delimiters.split(name)
    tokens = set()
    for t in raw_tokens:
        t = t.strip().lower()
        if not t:
            continue
        # 过滤纯数字或纯标点
        if re.match(r'^[\d\.\+\-\*×xX#]+$', t) and len(t) <= 6:
            continue
        tokens.add(t)
    return tokens


def compute_keyword_overlap(name_a, name_b):
    """
    计算关键词重叠系数（Overlap Coefficient）。
    overlap = |A ∩ B| / min(|A|, |B|)
    返回值 0.0 ~ 1.0
    """
    tokens_a = tokenize_name(name_a)
    tokens_b = tokenize_name(name_b)
    if not tokens_a or not tokens_b:
        return 0.0
    intersection = tokens_a & tokens_b
    return len(intersection) / min(len(tokens_a), len(tokens_b))


# 关键词重叠阈值：超过此值即判定为同一商品
OVERLAP_THRESHOLD = 0.6


def names_match(name_a, name_b):
    """
    判断两个名称是否指向同一商品（混合匹配策略）：
    1. 先尝试子串包含匹配（快速、精确）
    2. 子串不命中则用关键词重叠匹配（覆盖中间插入差异）
    任一命中即视为匹配。
    """
    na = normalize_name_for_compare(name_a)
    nb = normalize_name_for_compare(name_b)
    if not na or not nb:
        return False

    # 第一关：子串包含
    if na in nb or nb in na:
        return True

    # 第二关：关键词重叠
    overlap = compute_keyword_overlap(name_a, name_b)
    return overlap >= OVERLAP_THRESHOLD


def build_master_by_brand(ws, brand_col, name_col, code_col, header_row=1):
    """
    从商品原始表中按品牌分组构建索引。
    返回 (dict: brand → [(name, code), ...], rows_count)
    """
    by_brand = {}
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        brand = normalize_value(row[brand_col].value)
        name = normalize_value(row[name_col].value)
        code = normalize_value(row[code_col].value)
        if brand and name:
            by_brand.setdefault(brand, []).append((name, code))
            count += 1
    return by_brand, count


def check_and_mark(ws, master_by_brand, brand_col, name_col, code_col, header_row=1):
    """
    遍历动态商品池，检测「一品多码」：
    - 先按品牌匹配（精确优先，再子串包含兜底），筛选候选集
    - 名称匹配分三级：精确 > 子串包含 > 关键词重叠
    - 名称匹配且编码不同 → 一品多码，标红
    - 名称匹配且编码相同 → 同一商品，跳过
    返回 (duplicates_list, total_count, same_code_count, match_stats)
    """
    duplicates = []
    total = 0
    same_code_count = 0
    match_stats = {"exact": 0, "substring": 0, "overlap": 0}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        row_num = row[0].row
        pool_brand = normalize_value(row[brand_col].value)
        pool_name = normalize_value(row[name_col].value)
        pool_code = normalize_value(row[code_col].value) if code_col is not None else ""
        total += 1

        # 品牌匹配：精确优先，再子串包含兜底
        if pool_brand in master_by_brand:
            candidates = master_by_brand[pool_brand]
        else:
            # 子串包含：收集所有能匹配的品牌的候选列表
            candidates = []
            for master_brand, items in master_by_brand.items():
                if brands_match(pool_brand, master_brand):
                    candidates.extend(items)

        if not candidates:
            continue

        matched = False
        matched_master_name = ""
        matched_master_code = ""
        match_type = ""

        for master_name, master_code in candidates:
            if pool_name == master_name:
                matched = True
                matched_master_name = master_name
                matched_master_code = master_code
                match_type = "exact"
                break

        if not matched:
            for master_name, master_code in candidates:
                if names_match(pool_name, master_name):
                    matched = True
                    matched_master_name = master_name
                    matched_master_code = master_code
                    na = normalize_name_for_compare(pool_name)
                    nb = normalize_name_for_compare(master_name)
                    if na in nb or nb in na:
                        match_type = "substring"
                    else:
                        match_type = "overlap"
                    break

        if matched:
            match_stats[match_type] += 1
            if pool_code != matched_master_code:
                # 一品多码：名称匹配但编码不同
                duplicates.append((row_num, pool_brand, pool_name,
                                   pool_code, matched_master_code,
                                   matched_master_name, match_type))
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = RED_FILL
                        cell.font = WHITE_FONT
            else:
                same_code_count += 1

    return duplicates, total, same_code_count, match_stats


def validate_headers(ws, sheet_name, expected_cols):
    """校验表头是否与预期一致，返回列索引映射。"""
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    col_map = {}
    for field, expected_idx in expected_cols.items():
        if expected_idx >= len(headers):
            return None, f"Sheet「{sheet_name}」列数不足（期望至少{expected_idx + 1}列，实际{len(headers)}列）"
        actual = headers[expected_idx]
        if not actual:
            return None, f"Sheet「{sheet_name}」第{expected_idx + 1}列表头为空"
        col_map[field] = expected_idx
    return col_map, None


def main(file_path):
    """主流程"""
    # 1. 验证文件存在
    if not os.path.exists(file_path):
        print(json.dumps({"success": False, "error": f"文件不存在: {file_path}"}, ensure_ascii=False))
        sys.exit(1)

    # 2. 加载工作簿
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"无法打开文件: {str(e)}"}, ensure_ascii=False))
        sys.exit(1)

    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        print(json.dumps({
            "success": False,
            "error": f"数据总表需要至少2个Sheet，当前只有: {sheet_names}"
        }, ensure_ascii=False))
        wb.close()
        sys.exit(1)

    # 3. 获取两个Sheet（固定：第1个=原始表，第2个=商品池）
    master_sheet_name = sheet_names[0]
    pool_sheet_name = sheet_names[1]
    ws_master = wb[master_sheet_name]
    ws_pool = wb[pool_sheet_name]

    # 4. 校验表头
    master_map, err = validate_headers(ws_master, master_sheet_name, MASTER_COLS)
    if err:
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws_master[1]]
        print(json.dumps({
            "success": False,
            "error": f"{err}\n当前表头: {headers}\n期望: 商品编码 | 自定义名称 | 品牌"
        }, ensure_ascii=False))
        wb.close()
        sys.exit(1)

    pool_map, err = validate_headers(ws_pool, pool_sheet_name, POOL_COLS)
    if err:
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws_pool[1]]
        print(json.dumps({
            "success": False,
            "error": f"{err}\n当前表头: {headers}\n期望: 名称 | 品牌 | 编码 | 单位"
        }, ensure_ascii=False))
        wb.close()
        sys.exit(1)

    # 5. 构建原始表按品牌分组索引
    master_by_brand, master_count = build_master_by_brand(
        ws_master,
        master_map["品牌"],
        master_map["名称"],
        master_map["编码"],
    )

    # 6. 检查并标红
    duplicates, pool_count, same_code_count, match_stats = check_and_mark(
        ws_pool,
        master_by_brand,
        pool_map["品牌"],
        pool_map["名称"],
        pool_map["编码"],
    )

    # 7. 保存文件
    try:
        wb.save(file_path)
    except PermissionError:
        print(json.dumps({
            "success": False,
            "error": "无法保存文件，可能文件正在被其他程序（如Excel/WPS）打开。请关闭后再试。"
        }, ensure_ascii=False))
        wb.close()
        sys.exit(1)

    wb.close()

    # 8. 输出JSON结果
    result = {
        "success": True,
        "file_path": file_path,
        "master_sheet": master_sheet_name,
        "pool_sheet": pool_sheet_name,
        "master_headers": ["商品编码", "自定义名称", "品牌"],
        "pool_headers": ["名称", "品牌", "编码", "单位"],
        "master_count": master_count,
        "pool_count": pool_count,
        "duplicate_count": len(duplicates),
        "same_code_count": same_code_count,
        "match_stats": match_stats,
        "duplicates": [
            {
                "row": row_num,
                "brand": brand,
                "name": name,
                "pool_code": pool_code,
                "master_code": master_code,
                "master_name": master_name,
                "match_type": match_type,
            }
            for row_num, brand, name, pool_code, master_code, master_name, match_type in duplicates
        ],
    }

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "请提供数据总表文件路径"}, ensure_ascii=False))
        sys.exit(1)

    main(sys.argv[1])
